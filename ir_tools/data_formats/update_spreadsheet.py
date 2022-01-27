#!/usr/bin/env python

"""


Created: 
"""

import logging, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import xarray as xr
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

def modify_row_in_spreadsheet(path_fn, row_dict, col_names):
    shot_modify = row_dict['shot']

    # Convert units and types
    row_dict['date'] = datetime.datetime.strptime(row_dict['date'], '%Y-%m-%d')  # datetime.datetime()
    row_dict['exposure'] = row_dict['exposure'] * 1e-3  # convert us to ms
    row_dict['frame_period'] = row_dict['frame_period'] * 1e6  # convert s to us
    row_dict['detector_window'] = str(np.array(row_dict['detector_window'])[[1, 0, 3, 2]])   # Change order to [top, left, height, width]
    row_dict['n_frames'] = int(row_dict['n_frames'])

    row_values_new = list(row_dict.get(key, np.nan) for key in col_names)

    wb = load_workbook(path_fn)
    ws = wb.active

    header = True
    updated = False
    for i_row, row in enumerate(ws.iter_rows(), start=1):
        shot_cell = row[0].value
        row_values_old = list(cell.value for cell in row)

        if (not header) and (shot_cell is not None):
            shot = int(shot_cell)
            if shot < shot_modify:
                pass
            elif shot == shot_modify:
                logger.info(f'Updating existing row {i_row}: {row_values_old} -> {row_values_new}')
                for cell, value_new in zip(row, row_values_new):
                    if value_new is not np.nan:
                        cell.value = value_new
                updated = True
                break
            elif shot > shot_modify:
                logger.info(f'Inserting new row {i_row} for: {row_values_new}')
                ws.insert_rows(i_row, amount=1)
                row_new = ws[i_row]
                for cell, value_new in zip(row_new, row_values_new):
                    if value_new is not np.nan:
                        cell.value = value_new
                updated = True
                break


        if header and (shot_cell == 'Pulse'):
            header = False

    wb.save(path_fn)  # './meta_data_record/ss_modified.xlsx')


def modify_row_in_spreadsheet_old(path_fn, row_dict):
    shot_modify = row_dict['shot']
    row_values_new = list(row_dict.values())

    wb_old = load_workbook(path_fn)
    ws_old = wb_old.active
    wb_new = Workbook()
    ws_new = wb_new.create_sheet(ws_old.title)  # insert at the end (default)

    header = True
    for i_row, row in enumerate(ws_old.iter_rows()):
        shot_cell = row[0].value
        row_values_old = list(cell.value for cell in row)

        if header:
            ws_new.append(row_values_old)
        else:
            if shot_cell is not None:
                shot = int(shot_cell)
                if shot < shot_modify:
                    ws_new.append(row_values_old)
                elif shot == shot_modify:
                    # TODO: Merge old values
                    ws_new.append(row_values_new)
                elif shot > shot_modify:
                    ws_new.append(row_values_new)
                    ws_new.append(row_values_old)
            else:
                pass

        if header and (shot_cell == 'Pulse'):
            header = False

    wb_new.save('./meta_data_record/ss_modified.xlsx')

if __name__ == '__main__':
    path_fn = './meta_data_record/Record_of_MWIR1_FLIR_SC7500_6580045_Operating_Settings.xlsx'
    row_dict = {'shot': 33333}
    modify_row_in_spreadsheet(path_fn, row_dict=row_dict)
    pass